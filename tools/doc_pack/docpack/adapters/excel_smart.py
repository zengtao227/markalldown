from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from docpack.filters import format_numeric_summaries, render_markdown_table
from docpack.models import MissingDependencyError, PackOptions, PackResult, TableArtifact
from docpack.utils import slugify

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None


def build_pack(options: PackOptions) -> PackResult:
    if load_workbook is None:
        raise MissingDependencyError(
            "openpyxl is not installed. Install dependencies from tools/doc_pack/requirements.txt."
        )

    workbook = load_workbook(filename=str(options.source_path), read_only=True, data_only=True)
    sections = [f"# Spreadsheet Summary\n", f"Source: `{options.source_path.name}`\n"]
    tables: list[TableArtifact] = []
    metadata_sheets: list[dict[str, object]] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        artifact, sheet_markdown, sheet_metadata = _process_sheet(
            worksheet=worksheet,
            sheet_name=sheet_name,
            tables_dir=options.output_dir / "tables",
            preview_rows=options.preview_rows,
        )
        tables.append(artifact)
        metadata_sheets.append(sheet_metadata)
        sections.extend([sheet_markdown, ""])

    workbook.close()

    return PackResult(
        backend="excel-smart",
        document_type="spreadsheet",
        selected_units=options.selected_units or [],
        content_markdown="\n".join(section.strip() for section in sections if section.strip()),
        filters_applied=["table_summarize"],
        tables=tables,
        metadata={"sheet_count": len(metadata_sheets), "sheets": metadata_sheets},
    )


def _process_sheet(
    worksheet: object,
    sheet_name: str,
    tables_dir: Path,
    preview_rows: int,
) -> tuple[TableArtifact, str, dict[str, object]]:
    safe_name = slugify(sheet_name, fallback="sheet")
    csv_path = tables_dir / f"{safe_name}.csv"

    rows_iter = worksheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    header_values = list(header_row or [])
    column_count = max(worksheet.max_column or 0, len(header_values))
    if not header_values:
        header_values = [f"column_{index + 1}" for index in range(column_count)]
    headers = [
        _clean_header(value, index)
        for index, value in enumerate(header_values or [f"column_{index + 1}" for index in range(column_count)])
    ]

    row_count = 0
    preview: list[list[str]] = []
    numeric_stats: dict[str, dict[str, float]] = {}

    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)

        for row_index, row in enumerate(rows_iter, start=1):
            serialized = [_serialize_cell(value) for value in row]
            serialized += [""] * max(0, len(headers) - len(serialized))
            serialized = serialized[: len(headers)]
            writer.writerow(serialized)
            row_count += 1

            if row_index <= preview_rows:
                preview.append(serialized)

            for column_index, cell in enumerate(row):
                if column_index >= len(headers):
                    break
                number = _as_number(cell)
                if number is None:
                    continue
                stats = numeric_stats.setdefault(
                    headers[column_index],
                    {"count": 0.0, "sum": 0.0, "min": number, "max": number},
                )
                stats["count"] += 1
                stats["sum"] += number
                stats["min"] = min(stats["min"], number)
                stats["max"] = max(stats["max"], number)

    preview_markdown = render_markdown_table(headers, preview) if preview else "_No preview rows._"
    summary_lines = format_numeric_summaries(numeric_stats)
    summary_markdown = "\n".join(f"- {line}" for line in summary_lines) if summary_lines else "- None"

    markdown = "\n".join(
        [
            f"## Sheet: {sheet_name}",
            "",
            f"- CSV: `tables/{csv_path.name}`",
            f"- Size: {row_count} data rows x {len(headers)} columns",
            f"- Columns: {', '.join(f'`{header}`' for header in headers)}",
            "",
            "### Preview",
            "",
            preview_markdown,
            "",
            "### Numeric Summary",
            "",
            summary_markdown,
        ]
    )

    artifact = TableArtifact(
        path=csv_path,
        name=sheet_name,
        rows=row_count,
        columns=len(headers),
    )
    metadata = {
        "name": sheet_name,
        "csv": f"tables/{csv_path.name}",
        "rows": row_count,
        "columns": len(headers),
        "numeric_columns": sorted(numeric_stats.keys()),
    }
    return artifact, markdown, metadata


def _clean_header(value: object, index: int) -> str:
    text = _serialize_cell(value).strip()
    return text or f"column_{index + 1}"


def _serialize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _as_number(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None
